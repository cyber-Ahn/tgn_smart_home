## pip -> openpyxl==3.1.5
import shutil
import openpyxl 
from tgnLIB import *

phat_data = "/home/pi/tgn_smart_home/log/"
file_name = "room_data.log"

def build_exel(row_d,column_d,value_d):
    wb = openpyxl.load_workbook(phat_data+'Room_log.xlsx')
    ws = wb.active
    ws.cell(row=row_d, column=column_d, value = value_d)
    wb.save(phat_data+'Room_log.xlsx')

def clear_exel():
    wb = openpyxl.load_workbook(phat_data+'Room_log.xlsx')
    ws = wb.active
    ws.delete_rows(ws.min_row, ws.max_row)
    wb.save(phat_data+'Room_log.xlsx')

def convert_xlsx_to_pdf(xlsx_file):
    try:
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", xlsx_file])
        print("Done!")
    except Exception as e:
        print("Error:", e)

def build_log_file(phat_a):
    print("Build xlsx file on:" + phat_a)
    try:
        shutil.copyfile(phat_a+file_name, phat_a+"cach.tgn")
        f = open(phat_a+"cach.tgn", "r") 
        build_exel(1,1,"Date/Time")
        build_exel(1,2,"WZ Temp. C")
        build_exel(1,3,"WZ Hum. %")
        build_exel(1,4,"KZ Temp. C")
        build_exel(1,5,"KZ Hum. %")
        build_exel(1,6,"WC Temp. C")
        build_exel(1,7,"WC Hum. %")
        build_exel(1,8,"SZ Temp. C")
        build_exel(1,9,"SZ Hum. %")
        num_d = 2
        for line in f:
            line_date = line.rstrip().split("_")[0].split(".")[0]
            line_data = line.rstrip().split("_")[1]
            room_1 = line_data.split("|")[0].split(":")[1]
            room_2 = line_data.split("|")[1].split(":")[1]
            room_3 = line_data.split("|")[2].split(":")[1]
            room_4 = line_data.split("|")[3].split(":")[1]
            build_exel(num_d,1,line_date)
            build_exel(num_d,2,room_1.split("#")[0])
            build_exel(num_d,3,room_1.split("#")[1])
            build_exel(num_d,4,room_2.split("#")[0])
            build_exel(num_d,5,room_2.split("#")[1])
            build_exel(num_d,6,room_3.split("#")[0])
            build_exel(num_d,7,room_3.split("#")[1])
            build_exel(num_d,8,room_4.split("#")[0])
            build_exel(num_d,9,room_4.split("#")[1])
            num_d = num_d + 1
    except IOError:
        print("error...."+file_name+" file not found")
    time.sleep(2)
    print("Build pdf File")
    convert_xlsx_to_pdf(phat_data+"Room_log.xlsx")
    os.remove(phat_data+"cach.tgn")
    time.sleep(2)
    clear_exel()
    print("Log File ready in Room_log.pdf")
    open(phat_data+file_name, 'w').close()


build_log_file(phat_data)